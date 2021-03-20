# -*- coding: utf-8 -*-
"""
Created on Fri Mar 12 15:55:46 2021

@author: Lawrence
"""


stop=input('Press Enter to continue...')
try:
    import pandas as pd
    import numpy as np
    import time
    import os
    import datetime
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import utils
    from openpyxl.styles import Color, PatternFill, Font, Border,Alignment,Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule,Rule
    import warnings
    import math
    warnings.filterwarnings("ignore")
    # from sys import argv
    
    assgined_type_list=['Subcon','Buy & Sell','Fullbuy']
    # work_path=r'E:\desktop\3PM 自动化\二期需求\数据源-2'
    work_path=r'E:\desktop\lastest\数据源'
    try:
        most_updated_filename=input("Please input the latest FG plan filename...") #交付时放开注释
        last_updated_filename=input("Please input the compared FG plan filename...") #交付时放开注释
        # most_updated_filename='成品计划 2021 wk2.xlsx'
        # last_updated_filename='成品计划 2021 wk1.xlsx'
        most_updated_fg_plan=pd.read_excel(os.path.join(work_path,most_updated_filename),header=None)
        update_time_global=pd.to_datetime(most_updated_fg_plan.iloc[0,6])
        print('Current date is:',update_time_global.strftime("%Y/%m/%d"))
        print("Processing...")
    except:
        print('Most updated FG plan is invalid!')
    
    #-----------------------------MRP----------------------------------------------
    # try:
    #传入数据源
    def delete_word_warp(x):
        return [i.replace("\n",' ') for i in list(x)]
    #1 master data
    master_vendor=pd.read_excel(os.path.join(work_path,'Vendor Master Data.xlsx'))
    master_vendor.columns=delete_word_warp(master_vendor.columns)
    master_vendor['Vendor']=master_vendor['Vendor'].astype(str)
    master_vendor['FG']=master_vendor['FG'].astype(str)
    master_vendor.rename(columns={'Capacity (cs)':"Capacity (weekly)",
                                  "Unit of Capacity / MOQ":"MOQ Unit",
                                  },inplace=True)
    # master_vendor['MOQ(cs)']=master_vendor['Capacity (weekly)']
    master_vendor['MOQ Unit']='cs'
    master_vendor['Capacity Unit']='cs'
    
    master_vendor=master_vendor[['Vendor', 'Vendor Description', 'Mode', 'MOQ (cs)',
                                 'MOQ Unit','Capacity (weekly)', 'Capacity Unit',
                                 'LT', 'FG','FG Description']].rename(columns={"MOQ (cs)":"MOQ"})
    def convert_to_int(x):
        try:
            return int(x)
        except ValueError:
            pass
    master_vendor['MOQ']=master_vendor['MOQ'].apply(convert_to_int)
    master_vendor['Capacity (weekly)']=master_vendor['Capacity (weekly)'].apply(convert_to_int)
    
    #2 bom all data
    subcon=pd.read_excel(os.path.join(work_path,'BOM All.xlsx'),dtype={'FG':str,"RPM":str})
    subcon=subcon[['FG', 'FG Description', 'Base Quantity', 'Base Unit', 'RPM',
            'RPM Description', 'Material Type', 'Quantity with Scrap', 'Unit']]
    subcon.rename(columns={'RPM':'Material','RPM Description':'Material Description'},inplace=True)
    #3 成品未入库订单
    fg_to_be_delivered=pd.read_excel(os.path.join(work_path,'FG Open PO.XLSX'),dtype={'FG':str,"Vendor":str})
    fg_to_be_delivered.rename(columns={
            'PO Quantity':"To Be Delivered",
            "RDD":"Required Delivery Date",
            
        },inplace=True)
    #4 RPM未入库订单
    rpm_to_be_delivered=pd.read_excel(os.path.join(work_path,'RPM Open PO.xlsx'),dtype={'Supplier':str,"RPM":str})
    rpm_to_be_delivered=rpm_to_be_delivered[['RPM', 'RPM Description','Plant','Quantity','Unit',
                                           "PO","Item",'RDD','Supplier']]
    rpm_to_be_delivered.rename(columns={
        'RPM':"Material",
        'RPM Description':"Material Description",
        'Quantity':"To Be Delivered",
        "RDD":"Required Delivery Date",
        },inplace=True)
    #5 most updated fg plan
    raw_fg_plan=pd.read_excel(os.path.join(work_path,most_updated_filename),skiprows=2,dtype={'FG':str,"Vendor":str})
    #6 RPM帐内库存
    rpm_stock_in=pd.read_excel(os.path.join(work_path,'RPM Stock (账内).xlsx'),dtype={'Vendor':str,"RPM":str})
    rpm_stock_in.rename(columns={
            'RPM':"Material",
            'RPM Description':"Material Description",
        },inplace=True)
    #7 RPM master data
    master_rpm=pd.read_excel(os.path.join(work_path,'RPM Master Data.xlsx'),dtype={"Material":str})
    master_rpm=master_rpm[['Material',"LT",'Procurement Type','Material Type','MOQ']].rename(columns={
        'Material':u'物料',
        "LT":"PDT",
        "Procurement Type":'SPT',
        'MOQ':'RPM MOQ',
        })
    master_rpm.drop_duplicates(subset=u'物料',inplace=True)
    #8 FG priority
    fg_prio=pd.read_excel(os.path.join(work_path,'FG Priority.xlsx'),dtype={"FG":str})
    
    #------------------------------------------------------------------------------
    fg_rpm_dict=dict()
    for assgined_type in assgined_type_list:
        fg_plan=raw_fg_plan.copy()
        fg_plan=fg_plan.drop_duplicates(subset=['Vendor','FG'])
        fg_plan=fg_plan[(~pd.isna(fg_plan['Vendor']))&(~pd.isna(fg_plan['FG']))]
        fg_plan=fg_plan.fillna(0)
        fg_plan=pd.merge(fg_plan,master_vendor[['Vendor','FG','Mode']].drop_duplicates(subset=['Vendor','FG']),
                         on=['Vendor','FG'],how='left')
        # if assgined_type=='All':
        #     fg_plan=fg_plan[pd.notna(fg_plan['Mode'])].drop(columns='Mode')
        # elif assgined_type=='Subcon/Buy & Sell':
        #     fg_plan=fg_plan[(fg_plan['Mode']==assgined_type.split("/")[0])|(fg_plan['Mode']==assgined_type.split("/")[1])].drop(columns='Mode')
        # else:
        fg_plan=fg_plan[fg_plan['Mode']==assgined_type].drop(columns='Mode')
            
        #------------------------------------------------------------------------------
        new_col=[]
        cut_date=list(fg_plan.columns)[list(fg_plan.columns).index("Unit")+1]
        for i in list(fg_plan.columns):
            try:
                year=str(i.isocalendar()[0])
                week_num=str(i.isocalendar()[1])
                if len(week_num)<2:
                    week_num="0"+str(week_num)
                new_col.append(str(year)+"_wk"+str(week_num))
            except:
                new_col.append(i)
        fg_plan.columns=new_col
        
        fg_to_be_delivered['Required Delivery Date']=pd.to_datetime(fg_to_be_delivered['Required Delivery Date'])
        fg_to_be_delivered_2=fg_to_be_delivered[fg_to_be_delivered['Required Delivery Date']<cut_date]
        
        tem_df_1=fg_to_be_delivered_2.groupby(by=['Vendor','FG'],as_index=False).sum()[
            ['Vendor','FG','To Be Delivered']].rename(columns={'To Be Delivered':'Open PO'})
        
        fg_plan=pd.merge(fg_plan,tem_df_1,on=['Vendor','FG'],how='left')
        fg_plan['Open PO']=fg_plan['Open PO'].fillna(0)
        
        transform_fg_plan=fg_plan.melt(
            id_vars=list(fg_plan.columns)[:list(fg_plan.columns).index("Unit")+1],
            var_name='week num',
            value_name='FG fcst demand',
            )
        
        subcon_2=subcon.drop_duplicates(subset=["FG","Material"]).rename(columns={"Unit":'RPM Unit'})
        transform_fg_plan=pd.merge(
            transform_fg_plan,subcon_2[['FG','Base Quantity','Material','Material Description',
                                        'Quantity with Scrap','RPM Unit']],
            on='FG',how='inner'
            )
        
        transform_fg_plan['RPM fcst demand']=(transform_fg_plan['FG fcst demand']*transform_fg_plan['Quantity with Scrap']/transform_fg_plan['Base Quantity']).fillna(0)
        
        rpm_stock_in=rpm_stock_in[(~pd.isna(rpm_stock_in['Vendor']))&(~pd.isna(rpm_stock_in['Material']))]
        rpm_stock_in['RPM stock']=rpm_stock_in['Unrestricted'].fillna(0)+rpm_stock_in['In Quality Inspection'].fillna(0)
        transform_fg_plan_2=pd.merge(
            transform_fg_plan[transform_fg_plan['week num']=='Open PO'],
            rpm_stock_in.groupby(by=['Vendor','Material'],as_index=False).sum()[['Vendor','Material','RPM stock']],
            on=['Vendor','Material'],how='left'
            )
        transform_fg_plan_2['RPM stock']=transform_fg_plan_2['RPM stock'].fillna(0)
        
        def search_week_num(x):
            if x<cut_date:
                return "Open PO"
            else:
                year=str(x.isocalendar()[0])
                week_num=str(x.isocalendar()[1])
                if len(week_num)<2:
                    week_num="0"+str(week_num)
                return str(year)+"_wk"+str(week_num)
        
        rpm_to_be_delivered=rpm_to_be_delivered[(~pd.isna(rpm_to_be_delivered['Material']))&(~pd.isna(rpm_to_be_delivered['Required Delivery Date']))]
        rpm_to_be_delivered['Required Delivery Date']=pd.to_datetime(rpm_to_be_delivered['Required Delivery Date'])
        rpm_to_be_delivered['week num']=rpm_to_be_delivered['Required Delivery Date'].apply(search_week_num)
        
        transform_fg_plan_3=pd.concat([transform_fg_plan[transform_fg_plan['week num']!='Open PO'],
                                       transform_fg_plan_2],
                                      ignore_index=True)
        
        supply_summary=rpm_to_be_delivered.groupby(by=['Supplier','Material','week num'],as_index=False).sum()[
                                         ['Supplier','Material','week num','To Be Delivered']].rename(columns={'Supplier':'Vendor','To Be Delivered':'RPM supply'})
        
        initial_df=transform_fg_plan_3[transform_fg_plan_3['week num']=='Open PO']
        initial_df=pd.merge(initial_df,supply_summary,on=['Vendor','Material','week num'],how='left')
        initial_df['RPM supply']=initial_df['RPM supply'].fillna(0)
        # initial_df['begining stock in next week']=initial_df['RPM supply']+initial_df['RPM stock']-initial_df['RPM fcst demand']
        initial_df=pd.merge(initial_df,fg_prio[['FG','Priority Level']].drop_duplicates(),on='FG',how='left')
        allocate_stock=initial_df.groupby(by=['Vendor','Material'],as_index=False).min()[['Vendor','Material','Priority Level']].rename(columns={'Priority Level':'Highest'})
        initial_df=pd.merge(initial_df,allocate_stock,on=['Vendor','Material'],how='left')
        def allocate_stock_func(x):
            current_prio=x[0]
            highest=x[1]
            ava_stock=x[2]
            if current_prio>highest:
                return 0
            elif current_prio==highest:
                return ava_stock
        initial_df['RPM stock']=initial_df[['Priority Level','Highest','RPM stock']].apply(allocate_stock_func,axis=1)
        initial_df.drop(columns={'Priority Level','Highest'},inplace=True)
        
        initial_df['begining stock in next week']=initial_df['RPM stock']
        
        rolling_week=["Open PO"]
        rolling_week.extend(list(fg_plan.columns)[list(fg_plan.columns).index("Unit")+1:-1])
        
        result_df=initial_df.copy()
        
        for i in range(len(rolling_week)-1):
            initial_df=result_df[result_df['week num']==rolling_week[i]]
            df_this_week=transform_fg_plan_3[transform_fg_plan_3['week num']==rolling_week[i+1]]
            
            df_this_week=pd.merge(df_this_week,initial_df[['Vendor','FG','Material','begining stock in next week']],
                                  on=['Vendor','FG','Material'],how='left')
            df_this_week['RPM stock']=df_this_week['begining stock in next week']
            df_this_week['pre_cal_stock_in_next_week']=df_this_week['RPM stock']-df_this_week['RPM fcst demand']
            df_this_week=pd.merge(df_this_week,supply_summary,on=['Vendor','Material','week num'],how='left')
            df_this_week['RPM supply']=df_this_week['RPM supply'].fillna(0)
            
            fg_prio['Priority Level']=max(fg_prio['Priority Level'])+2-fg_prio['Priority Level']
            df_this_week=pd.merge(df_this_week,fg_prio[['FG','Priority Level']],on='FG',how='left')
            df_this_week['Priority Level']=df_this_week['Priority Level'].fillna(1)
            
            def prio_value_func(x):
                cal_stock=x[0]
                raw_prio=x[1]
                if cal_stock<0:
                    multiple=10000
                else:
                    multiple=1
                return raw_prio*multiple
            
            def allocate_supply(x):
                prio_value=x[0]
                max_prio_value=x[1]
                raw_supply=x[2]
                if prio_value==max_prio_value:
                    return raw_supply
                elif prio_value<max_prio_value:
                    return 0
                
            df_this_week['prio_value']=df_this_week[["pre_cal_stock_in_next_week","Priority Level"]].apply(
                prio_value_func,axis=1)
            df_this_week_2=pd.merge(df_this_week,
                                    df_this_week.groupby(by=['Vendor','Material'],as_index=False).max()[[
                                        'Vendor','Material',"prio_value"]].rename(columns={"prio_value":"max prio_value"}),
                                        how='left',on=['Vendor','Material'])
            # df_this_week_2['real RPM supply']=df_this_week_2[["prio_value","max prio_value",'RPM supply']].apply(allocate_supply,axis=1)
            df_this_week_2['RPM supply']=df_this_week_2[["prio_value","max prio_value",'RPM supply']].apply(allocate_supply,axis=1)
            # df_this_week_2['begining stock in next week']=df_this_week_2['RPM supply']+df_this_week_2['RPM stock']-df_this_week_2['RPM fcst demand']
            df_this_week_2['RPM stock']=df_this_week_2['RPM supply']+df_this_week_2['RPM stock']-df_this_week_2['RPM fcst demand']
            df_this_week_2['begining stock in next week']=df_this_week_2['RPM stock']
            result_df=pd.concat([result_df,df_this_week_2],ignore_index=True)
        
        transpose_df=result_df.drop(columns=['Base Quantity','Quantity with Scrap',
                                             'begining stock in next week','pre_cal_stock_in_next_week',
                                             'Priority Level','prio_value','max prio_value'])
        transpose_df=transpose_df.melt(
            id_vars=['Vendor', 'Vendor Description','FG','FG Description','Unit','week num',
                     'Material','Material Description','RPM Unit'],
            var_name='Type',
            value_name='value',
            )
        transpose_df['week num']=transpose_df['week num'].apply(lambda x: "1_Open PO" if x=="Open PO" else x)
        
        df_fg_plan=transpose_df[transpose_df['Type']=='FG fcst demand'].drop(columns='RPM Unit')
        df_fg_plan['Material']="00_"+df_fg_plan['FG']
        df_fg_plan['Material Description']=df_fg_plan['FG Description']
        
        df_rpm=transpose_df[transpose_df['Type']!='FG fcst demand']
        df_rpm['Unit']=df_rpm['RPM Unit']
        df_rpm.drop(columns='RPM Unit',inplace=True)
        
        df_fg_plan_2=df_fg_plan.drop_duplicates().set_index(keys=['Vendor', 'Vendor Description','FG','FG Description', 
                'Material', 'Material Description','Unit','Type', 'week num']).unstack()
        df_fg_plan_2.reset_index(drop=False,inplace=True)
        new_col=[]
        for i in list(df_fg_plan_2.columns):
            if i[1]:
                new_col.append(i[1])
            else:
                new_col.append(i[0])
        df_fg_plan_2.columns=new_col
        
        df_rpm_2=df_rpm.set_index(keys=['Vendor', 'Vendor Description','FG','FG Description', 
                'Material', 'Material Description','Unit','Type', 'week num']).unstack()
        df_rpm_2.reset_index(drop=False,inplace=True)
        new_col=[]
        for i in list(df_rpm_2.columns):
            if i[1]:
                new_col.append(i[1])
            else:
                new_col.append(i[0])
        df_rpm_2.columns=new_col
        
        result_df_2=pd.concat([df_fg_plan_2,df_rpm_2],ignore_index=True)
        def get_order(x):
            if x=='FG fcst demand':
                return "1_FG Plan"
            elif x=='RPM fcst demand':
                return "2_Demand"
            elif x=='RPM stock':
                return "3_Stock"
            elif x=='RPM supply':
                return "4_Supply"
            
        result_df_2['Type']=result_df_2['Type'].apply(get_order)
        result_df_2.sort_values(by=['Vendor','FG','Material','Type'],inplace=True)
        
        #-----------------------------------------------------------------------------
        vendor_material_comb=transpose_df[['Vendor','Material']].drop_duplicates().reset_index()
        vendor_material_list=[]
        for i in range(vendor_material_comb.shape[0]):
            vendor_material_list.append(list(vendor_material_comb.loc[i,:]))
        
        result_df_3=[]
        for index,vendor,material in vendor_material_list:
            selected_df=transpose_df[(transpose_df['Vendor']==vendor)&(transpose_df['Material']==material)]
            groupby_list=[]
            tem_df=selected_df[selected_df['Type']=='RPM fcst demand'].drop(columns=['Unit'])
            tem_df['Type']='sub RPM fcst demand'
            groupby_list.append(tem_df)
            for col_type in ['RPM fcst demand', 'RPM stock', 'RPM supply']:
            #     if col_type == 'FG fcst demand':
            #         tem_df=selected_df[selected_df['Type']==col_type].drop(columns=['RPM Unit','Unit'])
            #         groupby_list.append(tem_df)
                # else:
                tem_df=selected_df[selected_df['Type']==col_type]
                by_rpm=tem_df[tem_df['Type']==col_type].groupby(by=[
                    'Vendor', 'Vendor Description','Material', 'Material Description','week num','Type','RPM Unit'],as_index=False).sum()[
                        ['Vendor', 'Vendor Description','Material', 'Material Description','week num','Type','RPM Unit','value']]
                by_rpm['FG']=by_rpm['Material']
                by_rpm['FG Description']=by_rpm['Material Description']
                groupby_list.append(by_rpm)
            by_vendor_material_df=pd.concat(groupby_list,ignore_index=False)
            def get_order_2(x):
                if x=='sub RPM fcst demand':
                    return "2_Sub Demand"
                elif x=='RPM fcst demand':
                    return "1_Demand"
                elif x=='RPM stock':
                    return "3_Stock"
                elif x=='RPM supply':
                    return "4_Supply"
            by_vendor_material_df['Type']=by_vendor_material_df['Type'].apply(get_order_2)
            tem_result=by_vendor_material_df.set_index(keys=['Vendor', 'Vendor Description','Material',
                                                  'Material Description',"FG",'FG Description',"Type",'RPM Unit','week num']).unstack()
            
            tem_result.reset_index(drop=False,inplace=True)
            new_col=[]
            for i in list(tem_result.columns):
                if i[1]:
                    new_col.append(i[1])
                else:
                    new_col.append(i[0])
            tem_result.columns=new_col
            tem_result.sort_values(by=['Type'],inplace=True)
            result_df_3.append(tem_result)
            
        result_df_3=pd.concat(result_df_3,ignore_index=False)
        
        #-----------------------------------------------------------------------------
        date_col=list(result_df_2.columns)[list(result_df_2.columns).index("Type")+1:]
        result_df_2['Mode']=assgined_type
        result_df_2=pd.merge(result_df_2,
                             master_vendor[['Vendor','Mode','FG','Capacity (weekly)','MOQ','LT']].drop_duplicates(subset=['Vendor','Mode','FG']),
                             how='left',left_on=['Vendor','Mode','FG'],right_on=['Vendor','Mode','FG'])
        result_df_2=pd.merge(result_df_2,
                             master_rpm[[u'物料','RPM MOQ','Material Type']].drop_duplicates(subset=u'物料').rename(columns={u'物料':'Material'}),
                             how='left',on='Material')
        def moq_select(x):
            if x[0]=='1_FG Plan':
                return x[1]
            else:
                return x[2]
        result_df_2['MOQ']=result_df_2[['Type','MOQ','RPM MOQ']].apply(moq_select,axis=1)
        use_col=['Vendor', 'Vendor Description','Mode', 'FG', 'FG Description','Material',
               'Material Description','LT','MOQ','Capacity (weekly)','Unit','Type']
        use_col.extend(date_col)
        refine_df_2=result_df_2[use_col]
        
        result_df_3=pd.merge(result_df_3,
                             master_rpm[[u'物料','PDT','SPT','Material Type']].drop_duplicates(subset=u'物料').rename(columns={u'物料':'Material'}),
                             on='Material',how='left')
        use_col=['Vendor', 'Vendor Description','Material','Material Type','Material Description',
                 'FG','FG Description','SPT','PDT','RPM Unit','Type']
        use_col.extend(date_col)
        refine_df_3=result_df_3[use_col]
        
        #-----------------------------------------------------------------------------
        def split_str(x):
            if len(x.split("_"))==1:
                return x
            else:
                return x.split("_")[1]
        refine_df_2['Material']=refine_df_2['Material'].apply(split_str)
        refine_df_2['Type']=refine_df_2['Type'].apply(split_str)
        refine_df_2.rename(columns={'1_Open PO':'Original (Open PO)'},inplace=True)
        
        refine_df_3['Type']=refine_df_3['Type'].apply(split_str)
        refine_df_3.rename(columns={'1_Open PO':'Original (Open PO)',"RPM Unit":'Unit'},inplace=True)
        
        fg_rpm_dict[assgined_type]=[refine_df_2,refine_df_3,transpose_df]
    
    #This is the end of FG/RPM-----------------------------------------------------
    transpose_df_list=[]
    for key,value in fg_rpm_dict.items():
        transpose_df_list.append(value[2])
    transpose_df=pd.concat(transpose_df_list,ignore_index=True)
    
    def generate_one_dim_col(x):
        new_col=[]
        for i in list(x):
            if i[1]:
                new_col.append(i[1])
            else:
                new_col.append(i[0])
        return new_col
    pm_capacity=pd.read_excel(os.path.join(work_path,"PM Supplier Capacity.xlsx"),dtype={"Vendor":str,"PM":str})
    pm_capacity_comb=pm_capacity[['Vendor','PM','Capacity']].drop_duplicates().rename(columns={'PM':'Material'}).reset_index(drop=True)
    
    pm_capacity_comb_list=[]
    for i in range(pm_capacity_comb.shape[0]):
        pm_capacity_comb_list.append(list(pm_capacity_comb.iloc[i,:]))
    
    pm_capacity_result=[]
    for vendor,material,capacity in pm_capacity_comb_list:
        seleted_pm_df=transpose_df[
            (transpose_df['Vendor']==vendor)&(transpose_df['Material']==material)&
            (transpose_df['Type']=='RPM fcst demand')&(transpose_df['week num']!='1_Open PO')
            ]
        use_col=['Vendor', 'Vendor Description','Material', 'Material Description','FG',
                       'FG Description','RPM Unit','Type','week num','value']
        transpose_seleted_pm_df=seleted_pm_df[use_col].set_index(keys=use_col[:-1]).unstack()
        transpose_seleted_pm_df.reset_index(drop=False,inplace=True)
        transpose_seleted_pm_df.columns=generate_one_dim_col(transpose_seleted_pm_df.columns)
        transpose_seleted_pm_df['Type']='0_Sub Demand'
        
        
        group_by_col=['Vendor', 'Vendor Description','Material', 'Material Description',
                      'RPM Unit','week num','Type']
        group_seleted_pm_df=seleted_pm_df.groupby(by=group_by_col).sum()
        group_seleted_pm_df.reset_index(drop=False,inplace=True)
        group_seleted_pm_df['1_Demand']=group_seleted_pm_df['value']
        group_seleted_pm_df['2_Capacity']=capacity
        group_seleted_pm_df['3_Percentage']=group_seleted_pm_df['1_Demand']/capacity
        group_seleted_pm_df.drop(columns=['Type','value'],inplace=True)
        group_seleted_pm_df=group_seleted_pm_df.melt(
            id_vars=group_by_col[:-1],
            var_name='Type',
            value_name='value'
            )
        
        transpose_group_seleted=group_seleted_pm_df.set_index(keys=group_by_col).unstack(level=-2).reset_index(drop=False)
        transpose_group_seleted.columns=generate_one_dim_col(transpose_group_seleted.columns)
    
        sub_pm_capacity_result=pd.concat([transpose_seleted_pm_df,transpose_group_seleted],
                                         ignore_index=False)
        pm_capacity_result.append(sub_pm_capacity_result)
        
    pm_capacity_result=pd.concat(pm_capacity_result,ignore_index=False)
    # pm_capacity_result['FG'][pd.isna(pm_capacity_result['FG'])]=pm_capacity_result['Material']
    # pm_capacity_result['FG Description'][pd.isna(pm_capacity_result['FG Description'])]=pm_capacity_result['Material Description']
    
    pm_capacity_result['Type']=pm_capacity_result['Type'].apply(split_str)
    pm_capacity_result=pm_capacity_result.iloc[:,
                                               :list(pm_capacity_result.columns).index("Type")+14
                                               ]#仅保留前13周的数据
    #This is the end of PM Capacity Tracking---------------------------------------
    
    #7 4 RPM未入库订单
    rpm_to_be_delivered=pd.read_excel(os.path.join(work_path,'RPM Open PO.xlsx'),dtype={'Supplier':str,"RPM":str})
    rpm_to_be_delivered=rpm_to_be_delivered[['RPM', 'RPM Description','Plant','Quantity','Unit',
                                           "PO","Item",'RDD','Supplier','Docu Date']]
    rpm_to_be_delivered.rename(columns={
        'RPM':"Material",
        'RPM Description':"Material Description",
        'Quantity':"To Be Delivered",
        "RDD":"Required Delivery Date",
        'Supplier':'Vendor'
        },inplace=True)
    
    #7 RPM master data
    master_rpm=pd.read_excel(os.path.join(work_path,'RPM Master Data.xlsx'),dtype={"Material":str})
    master_rpm=master_rpm.rename(columns={
        "LT":"PDT",
        "Procurement Type":'SPT',
        'MOQ':'RPM MOQ',
        })
    master_rpm.drop_duplicates(subset='Material',inplace=True)
    
    transpose_df_list=[]
    for key,value in fg_rpm_dict.items():
        if key !='Fullbuy':
            transpose_df_list.append(value[2])
    transpose_df=pd.concat(transpose_df_list,ignore_index=True)
    
    use_col=['Vendor', 'Vendor Description','Plant','Storage Location','Material',
             'Material Description','Batch', 'Shelf Life Expired','Unit','Unrestricted',
             'In Quality Inspection', 'Restricted', 'Blocked']
    # rpm_stock_in_2=rpm_stock_in.groupby(by=use_col[:-4],as_index=False).sum()[use_col]
    rpm_stock_in_2=rpm_stock_in.copy()
    
    rpm_stock_in_2['Total']=rpm_stock_in_2[
        ['Unrestricted','In Quality Inspection','Restricted','Blocked']
        ].fillna(0).apply(lambda x: x.sum(),axis=1)
    rpm_stock_in_2['On Hand/On Order']='On Hand'
    rpm_stock_in_3=pd.merge(rpm_stock_in_2,master_rpm[['Material',"Material Type"]],on='Material',how='left')
    
    rpm_list=list(rpm_stock_in_2['Material'].unique())
    rpm_open_po_2=rpm_to_be_delivered[rpm_to_be_delivered['Material'].isin(rpm_list)]
    rpm_open_po_3=pd.merge(
        rpm_open_po_2.groupby(by=['Vendor','Plant','Material', 'Material Description','Unit','Docu Date'],
                              as_index=False).sum()[['Vendor','Plant','Material', 'Material Description','Unit','Docu Date','To Be Delivered']],
        master_rpm[['Material','Shelf Life',"Material Type"]],on='Material',how='left'
        )
    rpm_open_po_3.rename(columns={'To Be Delivered':'Total'},inplace=True)
    rpm_open_po_3['Shelf Life Expired']=rpm_open_po_3[['Docu Date',"Shelf Life"]].apply(
        lambda x : x[0]+datetime.timedelta(days=x[1]),axis=1
        )
    rpm_open_po_3['On Hand/On Order']='On Order'
    
    col_order=['Vendor', 'Vendor Description','Plant', 'Storage Location','Material','Material Description',"Material Type",
               'Batch','Shelf Life Expired', 'Unit','Unrestricted','In Quality Inspection',
               'Restricted', 'Blocked','Total','On Hand/On Order']
    
    stock_in_result=pd.concat([rpm_stock_in_3[col_order],
                               rpm_open_po_3[['Vendor', 'Plant', 'Material','Material Description',
                                             'Unit','Total','Material Type',
                                             'Shelf Life Expired', 'On Hand/On Order']]],
                               ignore_index=False)
    stock_in_result.sort_values(by=['Material','Vendor'],inplace=True)
    stock_in_result_2=pd.merge(stock_in_result,
        stock_in_result.groupby(by='Material',as_index=False).sum()[['Material','Total']].rename(columns={'Total':'By RPM Total'}),
        on='Material',how='left'
        )
    consumption_demand=transpose_df[
        (transpose_df['Type']=='RPM fcst demand') & (transpose_df['week num']!='1_Open PO')
        ].groupby(
        by='Material',as_index=False).sum()[['Material','value']].rename(columns={'value':'Forecast Consumption'})
    stock_in_result_2=pd.merge(stock_in_result_2,consumption_demand,on='Material',how='left')
    stock_in_result_2['Forecast Consumption']=stock_in_result_2['Forecast Consumption'].fillna(0)
    stock_in_result_2['Shelf Life Expired']=stock_in_result_2['Shelf Life Expired'].fillna(datetime.datetime(1900,1,1))
    def risk(x):
        expired=x[0]
        rpm_total=x[1]
        fcst_consump=x[2]
        if expired<=update_time_global+datetime.timedelta(days=90):
            return 'High'
        else:
            if rpm_total<fcst_consump:
                return 'Low'
            elif rpm_total>=fcst_consump:
                return 'High'
    stock_in_result_2['BW Risk']=stock_in_result_2[['Shelf Life Expired','By RPM Total',
                                                    'Forecast Consumption']].apply(risk,axis=1)
    # stock_in_result_2['Shelf Life Expired']=stock_in_result_2['Shelf Life Expired'].dt.strftime("%Y/%m/%d")
    stock_in_result_2.drop(columns='By RPM Total',inplace=True)
    
    #This is the end of stock(帐内)------------------------------------------------
    transpose_df=fg_rpm_dict['Fullbuy'][2]
    
    rpm_stock_out=pd.read_excel(os.path.join(work_path,'RPM Stock (账外).xlsx'),dtype={"Vendor":str,"RPM":str})
    rpm_stock_out_2=rpm_stock_out[['Vendor', 'Vendor Description', 'Material Type', 'RPM',
                                 'RPM Description', 'Batch', 'Shelf Life Expired', 'Unit',
                                 'Quantity']].rename(columns={'Quantity':'Total'})
    selected_df=transpose_df[
        (transpose_df['Type']=='FG fcst demand')&(transpose_df['week num']!='1_Open PO')
        ].groupby(by=['Vendor','Material'],as_index=False).sum()[['Vendor','Material','value']].rename(columns={'value':'FG Forecast Demand(cs)',"Material":'RPM'})
    
    rpm_stock_out_3=pd.merge(rpm_stock_out_2,selected_df,on=['Vendor','RPM'],how='left')
    rpm_stock_out_3['Shelf Life Expired']=pd.to_datetime(rpm_stock_out_3['Shelf Life Expired'])
    rpm_stock_out_3['BW Risk']=rpm_stock_out_3['Shelf Life Expired'].apply(lambda x: "High" if x<update_time_global else 'Low')
    # rpm_stock_out_3['Shelf Life Expired']=rpm_stock_out_3['Shelf Life Expired'].dt.strftime("%Y/%m/%d")
    
    #This is the end of stock(帐外)------------------------------------------------
    
    #cell style function
    def create_sheet(workbook,df,sheet_title):
        ws=workbook.create_sheet(title=sheet_title)
        max_col_index=utils.cell.get_column_letter(df.shape[1])
        max_row_index=str(df.shape[0]+1)
        font_object_content=Font(name=u'等线', bold=False, italic=False,size=10)
        font_object_title=Font(name=u'等线', bold=True, italic=False,size=11)
        alignment_title=Alignment(wrap_text=True)
        side = Side(style='thin',color='00000000')
        border = Border(left=side,right=side,top=side,bottom=side)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        max_col_index=utils.cell.get_column_letter(df.shape[1])
        max_row_index=str(df.shape[0]+1)
        for r in range(1,df.shape[0]+2): #all cell
            for c in range(1,df.shape[1]+1):
                cell_name=utils.cell.get_column_letter(c)+str(r)
                ws[cell_name].font=font_object_content
                ws[cell_name].border=border
                
        for c in range(1,df.shape[1]+1): #title cell
            cell_name=utils.cell.get_column_letter(c)+"1"
            ws[cell_name].font=font_object_title
            ws[cell_name].alignment=alignment_title
            ws[cell_name].fill=PatternFill(fill_type='solid', fgColor="00FABF8F") #补充标题的背景色
        return ws,max_col_index,max_row_index
    
    def row_style(ws,row_list,col_num,fgColor=False,font=False,num_format=False,start_col=0):
        for row in row_list:
            for col in range(1+start_col,1+col_num):
                target_cell= utils.cell.get_column_letter(col) + str((row+2))
                if fgColor:
                    ws[target_cell].fill=PatternFill(fill_type='solid', fgColor=fgColor)
                if font:
                    ws[target_cell].font=font
                if num_format:
                    ws[target_cell].number_format=num_format
        return ws
    
    wb_1 = Workbook()
    #------------------------------------------------------------------------------   
    for key,value in fg_rpm_dict.items():
        report_type=[' FG RPM',' RPM FG']
        for i in range(2):
            df=value[i].reset_index(drop=True)
            ws,max_col_index,max_row_index=create_sheet(wb_1,df,key+report_type[i])
            ws=row_style(ws,
                         row_list=df[df['Type']=='FG Plan'].index.tolist(),
                         col_num=df.shape[1],
                         font=Font(name=u'等线', bold=True, italic=False,size=10),
                         start_col=list(df.columns).index('Type'))
            ws=row_style(ws,
                         row_list=df[(df['Type']=='Demand') | (df['Type']=='Sub Demand')].index.tolist(),
                         col_num=df.shape[1],
                         fgColor="00FFF2CC",
                         start_col=list(df.columns).index('Type'))
            ws=row_style(ws,
                         row_list=df[df['Type']=='Stock'].index.tolist(),
                         col_num=df.shape[1],
                         fgColor="00DDEBF7",
                         start_col=list(df.columns).index('Type'))
            ws=row_style(ws,
                         row_list=df[df['Type']=='Supply'].index.tolist(),
                         col_num=df.shape[1],
                         fgColor="00E2EFDA",
                         start_col=list(df.columns).index('Type'))   
            ws=row_style(ws,
                         row_list=df[(df['Unit'].str.lower()=='cs') | (df['Unit'].str.lower()==u'件')].index.tolist(),
                         col_num=df.shape[1],
                         num_format='#,##0',
                         start_col=list(df.columns).index('Type'))   
            ws=row_style(ws,
                         row_list=df[(df['Unit'].str.lower()!='cs') & (df['Unit'].str.lower()!=u'件')].index.tolist(),
                         col_num=df.shape[1],
                         num_format='#,##0.000' ,
                         start_col=list(df.columns).index('Type')) 
            for i in df[df['Type']=='Stock'].index.tolist():
                for j in range(list(df.columns).index('Type')+1,df.shape[1]):
                    col=utils.cell.get_column_letter(j+1)
                    row=i+2
                    targrt_cell=col+str(row)
                    if float(df.iloc[i,j])<0:
                        ws[targrt_cell].font=Font(name=u'等线', bold=False, italic=False,size=10,color="00FF0000")
    #------------------------------------------------------------------------------        
    ws,max_col_index,max_row_index=create_sheet(wb_1,pm_capacity_result,'PM Capacity Tracking')
    df=pm_capacity_result.reset_index(drop=True)
    ws=row_style(ws,
                 row_list=df[(df['Type']=='Demand') | (df['Type']=='Sub Demand')].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00FFF2CC",
                 start_col=list(df.columns).index('Type'))
    ws=row_style(ws,
                 row_list=df[df['Type']=='Capacity'].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00C6E0B4",
                 start_col=list(df.columns).index('Type'))
    
    ws=row_style(ws,
                 row_list=df[(df['RPM Unit'].str.lower()=='cs') | (df['RPM Unit'].str.lower()==u'件')].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0',
                 start_col=list(df.columns).index('Type'))
    ws=row_style(ws,
                 row_list=df[df['Type']=='Percentage'].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00DDEBF7",
                 num_format='0%',
                 start_col=list(df.columns).index('Type'))
    
    for i in df[df['Type']=='Percentage'].index.tolist():
        for j in range(list(df.columns).index('Type')+1,df.shape[1]):
            col=utils.cell.get_column_letter(j+1)
            row=i+2
            targrt_cell=col+str(row)
            
            if 0.7<=float(df.iloc[i,j])<1:
                ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFC7CE")
            elif float(df.iloc[i,j])>=1:
                ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFEB9C")
    #------------------------------------------------------------------------------
    ws,max_col_index,max_row_index=create_sheet(wb_1,stock_in_result_2,'Stock 帐内')
    df=stock_in_result_2.reset_index(drop=True)
    ws=row_style(ws,
                 row_list=df[df['BW Risk']=='High'].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00FFC7CE",
                 start_col=list(df.columns).index('BW Risk'))
    ws=row_style(ws,
                 row_list=df[df['BW Risk']=='Low'].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00C6EFCE",
                 start_col=list(df.columns).index('BW Risk'))
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()=='cs') | (df['Unit'].str.lower()==u'件')].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0',
                 start_col=list(df.columns).index('Unit'))   
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()!='cs') & (df['Unit'].str.lower()!=u'件')].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0.000' ,
                 start_col=list(df.columns).index('Unit'))   
    #------------------------------------------------------------------------------
    ws,max_col_index,max_row_index=create_sheet(wb_1,rpm_stock_out_3,'Stock 帐外')
    df=rpm_stock_out_3.reset_index(drop=True)
    ws=row_style(ws,
                 row_list=df[df['BW Risk']=='High'].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00FFC7CE",
                 start_col=list(df.columns).index('BW Risk'))
    ws=row_style(ws,
                 row_list=df[df['BW Risk']=='Low'].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00C6EFCE",
                 start_col=list(df.columns).index('BW Risk'))
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()=='cs') | (df['Unit'].str.lower()==u'件')].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0',
                 start_col=list(df.columns).index('Unit'))   
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()!='cs') & (df['Unit'].str.lower()!=u'件')].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0.000',
                 start_col=list(df.columns).index('Unit'))   
    for i in range(df.shape[0]):
        col=utils.cell.get_column_letter(
            list(df.columns).index("FG Forecast Demand(cs)")+1)
        row=i+2
        target_cell=col+str(row)
        ws[target_cell].number_format='#,##0'
        
    wb_1.remove(wb_1['Sheet'])
    wb_1.save(os.path.join(work_path,"MRP.xlsx"))
    
    #-----------------------------Databse&Capacity Detail--------------------------
    # try:
    #去除表头的换行符
    def delete_word_warp(x):
        return [i.replace("\n",' ') for i in list(x)]
    
    #识别每个FG Plan的更新日期
    sub_path=[x for x in os.listdir(work_path) if (u'成品计划' in x)]
    return_result=[]
    for i in sub_path:
        tem_df=pd.read_excel(os.path.join(work_path,i),header=None)
        update_time=pd.to_datetime(tem_df.iloc[0,6]) #确定update date填在G1
        if len(str(update_time.isocalendar()[1]))==1:
            week_num='0'+str(update_time.isocalendar()[1])
        else:
            week_num=str(update_time.isocalendar()[1])
        version=str(update_time.isocalendar()[0]) +'_wk'+week_num
        return_result.append(
            list([os.path.join(work_path,i),version])
            )
    
    #Public mapping table for header transfomation
    mapping_table=pd.DataFrame(
        {'date':pd.date_range(start='2019/1/7', end='2030/12/31')}
        )
    mapping_table['week']=mapping_table['date'].dt.week
    mapping_table['year']=mapping_table['date'].dt.year
    mapping_table['weekday']=(mapping_table['date'].dt.weekday)+1
    mapping_table['corresponding Monday']=np.nan
    initial_Monday=mapping_table.loc[0,'date']
    for i in range(mapping_table.shape[0]):
        if mapping_table.loc[i,'weekday']==1:
            initial_Monday=mapping_table.loc[i,'date']
            mapping_table.loc[i,'corresponding Monday']=initial_Monday
        else:
            mapping_table.loc[i,'corresponding Monday']=initial_Monday
    mapping_table['standard week format']=mapping_table['year'].astype(str) +" wk" +mapping_table['week'].astype(str)
    mapping_table['string corresponding Monday']=mapping_table['corresponding Monday'].dt.strftime("%Y/%m/%d")
    mapping_table['string current date']=mapping_table['date'].dt.strftime("%Y/%m/%d")
    #reading raw datasource
    vendor_master=pd.read_excel(os.path.join(work_path,'Vendor Master Data.xlsx'),dtype={"Vendor":str,"FG":str})
    
    #combine all the FG Plan to into Sheet1 database
    com_df_list=[]
    for path,version in return_result:
        df_plan=pd.read_excel(os.path.join(work_path,path),skiprows=2,dtype={"Vendor":str,"FG":str})
        df_plan.insert(4,'Longest RPM LT',np.nan)
        df_plan.insert(6,'Version',version)
        com_df_list.append(df_plan)
    com_plan=pd.concat(com_df_list).sort_values(by=['Vendor',"FG",'Version'])
    #通过merge补充longest RPM LT信息
    com_plan=pd.merge(com_plan,vendor_master[['Vendor',"FG",'Longest\nRPM\nLT']],on=['Vendor',"FG"],how='left')
    com_plan['Longest RPM LT']=com_plan['Longest\nRPM\nLT'] 
    com_plan.drop(columns=['Longest\nRPM\nLT'],inplace=True) #用新列名取代原来表头
    #This is The end of Sheet1 Database____________________________________________
    
    #将竖表转置为横表，以plan version为列
    com_plan_viewer=pd.melt(com_plan,
        id_vars=['Vendor','Vendor Description','FG', 'FG Description','Longest RPM LT','Unit','Version'],
        var_name=['Date'])
    
    def convert_to_week(x):
        if len(str(x.isocalendar()[1]))==1:
            week_num='0'+str(x.isocalendar()[1])
        else:
            week_num=str(x.isocalendar()[1])
        return str(x.isocalendar()[0])+'_wk'+week_num
        
    # com_plan_viewer['Week']=com_plan_viewer['Date'].dt.year.astype(str) +' wk'+com_plan_viewer['Date'].dt.week.astype(str)
    com_plan_viewer['Week']=com_plan_viewer['Date'].apply(convert_to_week)
    
    com_plan_viewer.drop(columns=['Unit'],inplace=True)
    com_plan_viewer.drop_duplicates(subset=['Vendor','Vendor Description','FG', 'FG Description','Longest RPM LT','Week','Date','Version']
                                    ,inplace=True) #待确认FG plan是否有重复的行
    
    com_plan_viewer_2=com_plan_viewer.set_index(
        keys=['Vendor','Vendor Description','FG', 'FG Description','Longest RPM LT','Week','Date','Version'])
    
    com_plan_viewer_2=com_plan_viewer_2.unstack(level=-1).reset_index(drop=False) #将plan version转置到列
    new_col=[]
    for i in list(com_plan_viewer_2.columns):
        if i[1]=="":
            new_col.append(i[0])
        else:
            new_col.append(i[1])
    com_plan_viewer_2.columns=new_col #将multiple columns降维
    com_plan_viewer_2['Longest RPM LT'].fillna(0,inplace=True) #确认LT为空值如何处理
    com_plan_viewer_2.insert(7,'Status',np.nan)
    for i in range(com_plan_viewer_2.shape[0]): #进行是否block的判断
        if update_time_global + datetime.timedelta(days=(com_plan_viewer_2.loc[i,'Longest RPM LT']-7)) >= com_plan_viewer_2.loc[i,'Date'] :
            com_plan_viewer_2.loc[i,'Status']='Blocked'
        else:
            com_plan_viewer_2.loc[i,'Status']='Unblocked'
    viewer_value_header=list(com_plan_viewer_2)[-len(sub_path):]
    #填充FG PLAN中的空值
    dummy_date=pd.DataFrame(pd.date_range(start='2010/1/1',end='2030/12/31'),columns={'date'})
    dummy_date['week_num']=dummy_date['date'].apply(convert_to_week)
    week_order=list(dummy_date['week_num'].unique())
    com_plan_viewer_2.reset_index(drop=True,inplace=True)
    for col in list(com_plan_viewer_2.columns)[-len(sub_path)+1:]:
        for i in range(com_plan_viewer_2.shape[0]):
            left_order=week_order.index(str(com_plan_viewer_2.loc[i,'Week']))
            header_order=week_order.index(col)
            if left_order<=header_order:
                com_plan_viewer_2.loc[i,col]=com_plan_viewer_2.iloc[i,list(com_plan_viewer_2.columns).index(col)-1]
    com_plan_viewer_2.loc[:,viewer_value_header]=com_plan_viewer_2.loc[:,viewer_value_header].fillna(0)
    
    com_plan_viewer_3=com_plan_viewer_2.copy()
    #计算相邻两份version预测的差值
    for i in range(1,len(sub_path)):
        current_period=viewer_value_header[i]
        last_period=viewer_value_header[i-1]
        new_col_name=current_period+'\n-\n'+last_period
        com_plan_viewer_3[new_col_name]=com_plan_viewer_3[current_period]-com_plan_viewer_3[last_period]
    #按照Vnedor-FG-Date的顺序进行排列
    com_plan_viewer_3.sort_values(by=['Vendor',"FG",'Date'],inplace=True)
    com_plan_viewer_3['Date']=com_plan_viewer_3['Date'].dt.strftime("%Y/%m/%d")
    #This is The end of Sheet2 Database Viewer_____________________________________
    
    #从vendor master data中提取basci info
    vendor_master.columns=delete_word_warp(vendor_master.columns)
    capacity_df=vendor_master[['Planner', 'Vendor', 'Vendor Description','Working Day /  Wk',
                               'Working Hour /  Day', 'Mode', 'Production Type', 'Classification',
                               'No. of Lines','Delicated or Shared','FG', 'FG Description',
                               'Unit of Capacity / MOQ','Capacity', 'MOQ','If Delist? ']]
    
    capacity_df=capacity_df[
        (pd.notna(capacity_df['Classification'])) & (pd.isna(capacity_df['If Delist? ']))
        ] #暂时filter掉分类为空的行  & 去除再delist中的记录
    capacity_df.drop(columns='If Delist? ',inplace=True)
    
    fg_master=pd.read_excel(os.path.join(work_path,'FG Master Data.XLSX'),skiprows=1)
    fg_master.columns=delete_word_warp(fg_master.columns)
    fg_master.rename(columns={"Unnamed: 0":"FG","Unnamed: 1":"FG Description", u"箱规":'Pack/CS', 'NW':'NW/KG'},inplace=True)
    fg_master['FG']=fg_master['FG'].astype(str)
    fg_master=fg_master.drop_duplicates(subset='FG')
    
    # fg_master['LT']=0 #辅助的列，后期确定后需要删去
    fg_master=pd.merge(
        fg_master,vendor_master[['FG','LT']].drop_duplicates(),on='FG',how='left'
        )
    #利用FG master补全basci info
    capacity_df=pd.merge(capacity_df,fg_master[["FG",'Pack/CS','NW/KG',"LT"]],on='FG',how='left')
    #most updated的确定方式待确认
    most_updated_fg_plan=pd.read_excel(os.path.join(work_path,most_updated_filename),skiprows=2,dtype={"Vendor":str,"FG":str})
    
    #转换表头的时间类型，转为字符
    new_col=[]
    for i in list(most_updated_fg_plan.columns):
        try:
            new_i=i.strftime("%Y/%m/%d")
        except:
            new_i=i
        new_col.append(new_i)
    most_updated_fg_plan.columns=new_col
    
    keep_col=['Vendor',"FG",'Unit']
    keep_col.extend(
        list(most_updated_fg_plan)[list(most_updated_fg_plan).index("Unit")+1:]
        ) #仅保留预测值的部分和merge所需的列
    
    capacity_df_2=capacity_df.copy()
    capacity_df_2['Info Type']="Plan"
    
    #merge后行数可能变多
    capacity_df_2=pd.merge(capacity_df_2,most_updated_fg_plan[keep_col].drop_duplicates(subset=['Vendor',"FG"]),on=['Vendor',"FG"],how='left')  #确认是否保留mapping不到的Vendor-FG 
    capacity_df_2[keep_col[3:]]=capacity_df_2[keep_col[3:]].fillna(0) #补充预测值中的空值
    
    #为by vendor-classification处理做好准备，产生唯一的组合进行后续的迭代
    vendor_class_comb=capacity_df_2[['Vendor','Classification']].drop_duplicates()
    vendor_class_comb_list=[]
    for i in range(vendor_class_comb.shape[0]):
        vendor_class_comb_list.append(
            tuple(vendor_class_comb.iloc[i,:])
            )
    
    #by vendor/classifiaction对split df进行处理，之后合并到capacity detail中
    capacity_detail=[]
    for vendor,classification in vendor_class_comb_list:
        split_df=capacity_df_2[
            (capacity_df_2['Vendor']==vendor) & (capacity_df_2["Classification"]==classification)
            ].reset_index(drop=True)
        append_row=dict()
        for i in ['Planner','Vendor', 'Vendor Description','Working Day /  Wk','Working Hour /  Day',
                  'Mode','Production Type', 'Classification','No. of Lines',
                  'Delicated or Shared','Unit of Capacity / MOQ','Capacity']:
            append_row[i]=split_df.loc[0,i]
        append_row_1=append_row.copy()
        append_row_2=append_row.copy()
        append_row_1['Info Type']=u'总量'
        append_row_2[u'Info Type']=u'产能利用率'
        capacity_unit=str(split_df.loc[0,'Unit of Capacity / MOQ']).lower()
        
        if pd.notna(split_df.loc[0,'Capacity']):
            capacity_volume=float(split_df.loc[0,'Capacity'])
            for j in range(1,len(keep_col)-2):
                if capacity_unit=='ton':
                    append_row_1[
                        list(split_df.columns)[-j]
                        ]=(split_df.iloc[:,-j]*split_df.loc[:,"NW/KG"]).sum()/1000
                    append_row_2[
                        list(split_df.columns)[-j]
                        ]=(split_df.iloc[:,-j]*split_df.loc[:,"NW/KG"]).sum()/(1000*capacity_volume)         
                elif capacity_unit=='pcs' or capacity_unit=='pck':
                    append_row_1[
                        list(split_df.columns)[-j]
                        ]=((split_df.iloc[:,-j]*split_df.loc[:,"NW/KG"]).sum())
                    append_row_2[
                        list(split_df.columns)[-j]
                        ]=((split_df.iloc[:,-j]*split_df.loc[:,"Pack/CS"]).sum())/capacity_volume
                
                append_row_1['Unit']=capacity_unit
                append_row_2['Unit']="fill rate"
        
            split_df=split_df.append(append_row_1,ignore_index=True)
            split_df=split_df.append(append_row_2,ignore_index=True)
            
            capacity_detail.append(split_df)
    capacity_detail=pd.concat(capacity_detail) #此处完成Plan和fill rate的计算
    
    #reading raw datasource
    fg_open_po=pd.read_excel(os.path.join(work_path,"FG Open PO.XLSX"),dtype={"Vendor":str,"FG":str})
    fg_open_po=pd.merge(fg_open_po,mapping_table[['date','string corresponding Monday']],left_on='RDD',
                        right_on='date',how='left')
    
    #借用plan类的部分，对预测部分进行改写
    capacity_detail_po=capacity_detail[capacity_detail['Info Type']=='Plan'].copy()
    start_col_num=list(capacity_detail.columns).index("Unit")+1
    capacity_detail_po.iloc[:,start_col_num:]=np.nan #先将原来的预测清空
    capacity_detail_po['Info Type']='PO' #重写类型
    
    capacity_detail_po.reset_index(drop=True,inplace=True) #为后面的shape循环做好准备
    for i in range(capacity_detail_po.shape[0]):
        for j in range(start_col_num,capacity_detail_po.shape[1]):
            index_tuple=tuple([
                capacity_detail_po.loc[i,'Vendor'],
                capacity_detail_po.loc[i,'FG'],
                list(capacity_detail_po.columns)[j]
                ])
            
            tem_df=fg_open_po[
                (fg_open_po['Vendor']==index_tuple[0])&
                (fg_open_po['FG']==index_tuple[1])&
                (fg_open_po['string corresponding Monday']==index_tuple[2])
                ]
            
            if tem_df['PO Quantity'].sum() > 0: #如果存在open po则填入总量，否则留空
                capacity_detail_po.iloc[i,j]=tem_df['PO Quantity'].sum()
            
    earliest_week=pd.to_datetime(list(most_updated_fg_plan.columns)[5],format="%Y/%m/%d") #根据most updated plan来定
     
    fg_open_po_before=fg_open_po[fg_open_po['date']<earliest_week]
    fg_open_po_before=fg_open_po_before.rename(columns={"PO Quantity":"Open PO"})    
    capacity_detail_po=pd.merge(
        capacity_detail_po,
        fg_open_po_before.groupby(by=["Vendor","FG"],as_index=False).sum()[["Vendor","FG","Open PO"]],
        on=["Vendor","FG"],how='left'
        ) #将open po合并到 po detail中
    
    capacity_detail.insert(
        list(capacity_detail.columns).index("Unit")+1,"Open PO",np.nan
        ) #在plan类表中添加一个表头，便于合并后调整表头顺序
    
    comb_capacity_detail=pd.concat([capacity_detail,capacity_detail_po]).reset_index(drop=True)
    order_table=pd.DataFrame(
        {"Info Type":["Plan","PO",u"总量",u"产能利用率"],
         "Dummy":["1-Plan","2-PO",u"3-总量",u"4-产能利用率"]
         }
        )
    comb_capacity_detail=pd.merge(
        comb_capacity_detail,order_table,on='Info Type',how='left'
        )
    comb_capacity_detail.sort_values(by=['Vendor',"Classification","FG","Dummy"],inplace=True)
    comb_capacity_detail.drop(columns='Dummy',inplace=True)
    comb_capacity_detail=comb_capacity_detail.iloc[:,
                                                   :list(comb_capacity_detail.columns).index("Open PO")+14
                                                   ]#仅保留前13周数据
    
    # dummy_df=comb_capacity_detail.reset_index(drop=True)
    # red_list_comb_capacity_detail=[]
    # yellow_list_comb_capacity_detail=[]
    # mild_yellow_list_comb_capacity_detail=[]
    
    # for i in range(dummy_df.shape[0]):
    #     if dummy_df.loc[i,"Info Type"]==u'产能利用率':
    #         for j in range(list(dummy_df.columns).index("Open PO")+1,dummy_df.shape[1]):
    #             if float(dummy_df.iloc[i,j])>1:
    #                 red_list_comb_capacity_detail.append([(i),(j)])
    #             elif 0.7<=float(dummy_df.iloc[i,j])<=1:
    #                 yellow_list_comb_capacity_detail.append([(i),(j)])
                    
    comb_capacity_detail.reset_index(inplace=True,drop=True)
    for i in range(comb_capacity_detail.shape[0]):
        if str(comb_capacity_detail.loc[i,"Unit"])=='nan' and str(comb_capacity_detail.loc[i,"Info Type"])==u"总量":
            comb_capacity_detail.loc[i,"Unit"]="" #填充unit为空的列为
        if str(comb_capacity_detail.loc[i,"Info Type"])=="Plan"  or str(comb_capacity_detail.loc[i,"Info Type"])=="PO":
            comb_capacity_detail.loc[i,"Unit"]="CS"
    
    # for i in range(comb_capacity_detail.shape[0]): #标记current date+LT 需要bloked的行
    #     if str(comb_capacity_detail.loc[i,"Info Type"])=="PO" and pd.notnull(comb_capacity_detail.loc[i,"LT"]):
    #         week_num=math.floor(float(comb_capacity_detail.loc[i,"LT"])/7)
    #         if week_num>0:
    #             for j in range(week_num):
    #                 initial_value=list(comb_capacity_detail.columns).index("Open PO")+1
    #                 mild_yellow_list_comb_capacity_detail.append([(i),(j+initial_value)])
    
    #补充comment列需求
    # comb_capacity_detail['Comment']=np.nan
    current_fg_plan=pd.read_excel(os.path.join(work_path,most_updated_filename),skiprows=2,dtype={'FG':str,"RPM":str})
    last_fg_plan=pd.read_excel(os.path.join(work_path,last_updated_filename),skiprows=2,dtype={'FG':str,"RPM":str})
    def delta_f(x):
        if x>0:
            return u'增量'+format(abs(x),'.0f')
        elif x==0:
            return u'无变化'
        elif x<0:
            return u'减量'+format(abs(x),'.0f')
    result_df=[]
    compare_col=list(current_fg_plan.columns)[
        list(current_fg_plan.columns).index("Unit")+1:list(current_fg_plan.columns).index("Unit")+14]
    for i in compare_col:
        if i in list(last_fg_plan):
            keep_col=['Vendor',"FG"]
            keep_col.append(i)
            current_df=current_fg_plan[keep_col].rename(columns={i:"current"+' wk-'+str(i.isocalendar()[1])})
            last_df=last_fg_plan[keep_col].rename(columns={i:"last"+' wk-'+str(i.isocalendar()[1])})
            com_df=pd.merge(current_df,last_df,on=['Vendor',"FG"],how='left')
            com_df.fillna(0,inplace=True)
            com_df['delta']=com_df.iloc[:,-2]-com_df.iloc[:,-1]
            com_df['sub comment']=com_df['delta'].apply(delta_f)
            com_df['week']='wk'+str(i.isocalendar()[1])
            com_df['comple comment']=com_df[['week','sub comment']].apply(lambda x:str(x[0])+str(x[1]),axis=1)
            # com_df=com_df.iloc[:,[0,1,5,6]]
            com_df=com_df.iloc[:,[0,1,7]]
            result_df.append(com_df)
            
    result_df=pd.concat(result_df).reset_index(drop=True)
    vendor_fg_combine=result_df[['Vendor',"FG"]].drop_duplicates().reset_index(drop=True)
    vendor_fg_combine['Comment']=np.nan
    for i in range(vendor_fg_combine.shape[0]):
        tem_list=list(vendor_fg_combine.iloc[i,])
        tem_result_df=result_df[(result_df['Vendor']==tem_list[0]) & (result_df['FG']==tem_list[1])].drop_duplicates()
        comple_comment=",".join(list(tem_result_df['comple comment']))
        vendor_fg_combine.loc[i,'Comment']=comple_comment
    vendor_fg_combine['Vendor']=vendor_fg_combine['Vendor'].astype(str)
    comb_capacity_detail=pd.merge(comb_capacity_detail,vendor_fg_combine,on=['Vendor',"FG"],how='left')
    #去除非plan列的comment
    for i in range(comb_capacity_detail.shape[0]):
        if comb_capacity_detail.loc[i,'Info Type']!='Plan':
            comb_capacity_detail.loc[i,'Comment']=np.nan
    #This is The end of Sheet3 Capacity Detail_____________________________________
    
    summary_detail=comb_capacity_detail[
        (comb_capacity_detail['Info Type']==u'总量') | (comb_capacity_detail['Info Type']==u'产能利用率')]
    summary_detail.drop(columns='Comment',inplace=True)
    # red_list_summary_detail=[]
    # yellow_list_summary_detail=[]
    # dummy_df=summary_detail.reset_index(drop=True)
    # for i in range(dummy_df.shape[0]):
    #     if dummy_df.loc[i,"Info Type"]==u'产能利用率':
    #         for j in range(list(dummy_df.columns).index("Open PO")+1,dummy_df.shape[1]):
    #             if float(dummy_df.iloc[i,j])>1:
    #                 red_list_summary_detail.append([(i),(j)])
    #             elif 0.7<=float(dummy_df.iloc[i,j])<=1:
    #                 yellow_list_summary_detail.append([(i),(j)])
    summary_detail.reset_index(inplace=True,drop=True)
    #This is The end of Sheet4 Capacity Summary____________________________________
    old_col=list(com_plan.columns)
    new_col=[]
    for i in old_col:
        try:
            new_col.append(i.strftime("%Y/%m/%d"))
        except:
           new_col.append(i) 
    com_plan.columns=new_col
    
    #cell style function
    def create_sheet(workbook,df,sheet_title):
        ws=workbook.create_sheet(title=sheet_title)
        max_col_index=utils.cell.get_column_letter(df.shape[1])
        max_row_index=str(df.shape[0]+1)
        font_object_content=Font(name=u'等线', bold=False, italic=False,size=10)
        font_object_title=Font(name=u'等线', bold=True, italic=False,size=11)
        alignment_title=Alignment(wrap_text=True)
        side = Side(style='thin',color='00000000')
        border = Border(left=side,right=side,top=side,bottom=side)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        max_col_index=utils.cell.get_column_letter(df.shape[1])
        max_row_index=str(df.shape[0]+1)
        for r in range(1,df.shape[0]+2): #all cell
            for c in range(1,df.shape[1]+1):
                cell_name=utils.cell.get_column_letter(c)+str(r)
                ws[cell_name].font=font_object_content
                ws[cell_name].border=border
                
        for c in range(1,df.shape[1]+1): #title cell
            cell_name=utils.cell.get_column_letter(c)+"1"
            ws[cell_name].font=font_object_title
            ws[cell_name].alignment=alignment_title
            ws[cell_name].fill=PatternFill(fill_type='solid', fgColor="00FABF8F") #补充标题的背景色
        return ws,max_col_index,max_row_index
    
    def row_style(ws,row_list,col_num,fgColor=False,font=False,num_format=False,start_col=0):
        for row in row_list:
            for col in range(1+start_col,1+col_num):
                target_cell= utils.cell.get_column_letter(col) + str((row+2))
                if fgColor:
                    ws[target_cell].fill=PatternFill(fill_type='solid', fgColor=fgColor)
                if font:
                    ws[target_cell].font=font
                if num_format:
                    ws[target_cell].number_format=num_format
        return ws
    
    
    wb_1 = Workbook()
    #------------------------------------------------------------------------------
    ws,max_col_index,max_row_index=create_sheet(wb_1,com_plan,"Database")
    for col in range(list(com_plan.columns).index("Version")+1,com_plan.shape[1]+1):
        for row in range(com_plan.shape[0]):
            target_cell= utils.cell.get_column_letter(col) + str((row+2))
            ws[target_cell].number_format= '#,##0'
    #------------------------------------------------------------------------------      
    ws,max_col_index,max_row_index=create_sheet(wb_1,com_plan_viewer_3,"Data Base Viewer")
    
    block_fill = PatternFill(bgColor="00FFD966") #fill cell if blocked
    dxf = DifferentialStyle(fill=block_fill)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = ['$H1="Blocked"']
    ws.conditional_formatting.add("H1:"+("H"+max_row_index),r)
    
    unblock_fill = PatternFill(bgColor="00A9D08E") #fill cell if unblocked
    dxf = DifferentialStyle(fill=unblock_fill)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = ['$H1="Unblocked"']
    ws.conditional_formatting.add("H1:"+("H"+max_row_index),r)
    
    df=com_plan_viewer_3.reset_index(drop=True)
    for col in range(list(df.columns).index("Status")+1,df.shape[1]+1):
        for row in range(df.shape[0]):
            target_cell= utils.cell.get_column_letter(col) + str((row+2))
            ws[target_cell].number_format= '#,##0'
            
    target_col=list(com_plan_viewer_3.columns)[-1]
    ws=row_style(ws,
                 row_list=df[df[target_col]!=0].index.tolist(),
                 col_num=df.shape[1],
                 fgColor="00FFC7CE",
                 start_col=list(df.columns).index(target_col))
    #------------------------------------------------------------------------------  
    ws,max_col_index,max_row_index=create_sheet(wb_1,comb_capacity_detail,"Capacity Detail")
    max_range=max_col_index+max_row_index
    
    df=comb_capacity_detail.reset_index(drop=True) 
    ws=row_style(ws,
                  row_list=df[df['Info Type']==u'总量'].index.tolist(),
                  col_num=df.shape[1]-1, #不包含comment列
                  fgColor="00E2EFDA")
    ws=row_style(ws,
                  row_list=df[df['Info Type']==u'产能利用率'].index.tolist(),
                  col_num=df.shape[1]-1,
                  fgColor="00C6E0B4")
    ws=row_style(ws,
                  row_list=df[df['Info Type']==u'PO'].index.tolist(),
                  col_num=df.shape[1]-1,
                  fgColor="00DDEBF7")
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()=='cs') | (df['Unit'].str.lower()=='pcs') | (df['Unit'].str.lower()==u'件') | (df['Unit'].str.lower()=='pck') ].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0' ,
                 start_col=list(df.columns).index('Unit'))   
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()!='cs') & (df['Unit'].str.lower()!='pcs') & (df['Unit'].str.lower()!=u'件') & (df['Unit'].str.lower()!='pck')].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0.000' ,
                 start_col=list(df.columns).index('Unit'))   
    
    for i in df[df['Info Type']==u'产能利用率'].index.tolist():
        for j in range(list(df.columns).index('Open PO')+1,df.shape[1]-1):
            col=utils.cell.get_column_letter(j+1)
            row=i+2
            targrt_cell=col+str(row)
            ws[targrt_cell].number_format="0%"
            if 0.7<=float(df.iloc[i,j])<1:
                ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFEB9C")
            elif float(df.iloc[i,j])>=1:
                ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFC7CE")
                
    for i in df[df['Info Type']==u'PO'].index.tolist():
        try:
            forward_step=math.floor(float(df.loc[i,'LT'])/7)
            if forward_step>0:
                for j in range(list(df.columns).index('Open PO')+1,list(df.columns).index('Open PO')+1+forward_step):
                    col=utils.cell.get_column_letter(j+1)
                    row=i+2  
                    targrt_cell=col+str(row)
                    ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFFFA7")
        except ValueError:
            pass
    #------------------------------------------------------------------------------
    ws,max_col_index,max_row_index=create_sheet(wb_1,summary_detail,"Capacity Summary")
    max_range=max_col_index+max_row_index
    
    df=summary_detail.reset_index(drop=True) 
    
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()=='cs') | (df['Unit'].str.lower()=='pcs') | (df['Unit'].str.lower()==u'件') ].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0' ,
                 start_col=list(df.columns).index('Unit'))   
    ws=row_style(ws,
                 row_list=df[(df['Unit'].str.lower()!='cs') & (df['Unit'].str.lower()!='pcs') & (df['Unit'].str.lower()!=u'件') ].index.tolist(),
                 col_num=df.shape[1],
                 num_format='#,##0.000' ,
                 start_col=list(df.columns).index('Unit'))   
    
    ws=row_style(ws,
                  row_list=df[df['Info Type']==u'产能利用率'].index.tolist(),
                  col_num=df.shape[1],
                  fgColor="00C6E0B4")
    for i in df[df['Info Type']==u'产能利用率'].index.tolist():
        for j in range(list(df.columns).index('Open PO')+1,df.shape[1]):
            col=utils.cell.get_column_letter(j+1)
            row=i+2
            targrt_cell=col+str(row)
            ws[targrt_cell].number_format="0%"
            if 0.7<=float(df.iloc[i,j])<1:
                ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFEB9C")
            elif float(df.iloc[i,j])>=1:
                ws[targrt_cell].fill=PatternFill(fill_type='solid', fgColor="00FFC7CE")
                
    
    
    wb_1.remove(wb_1['Sheet'])
    wb_1.save(os.path.join(work_path,"Capacity&Data Base.xlsx"))


except Exception as e:
    print('\nOoops,somthing goes wrong...Here is the information about the error:')
    print(repr(e))

stop=input("\nComplete,press any key to exit...")
